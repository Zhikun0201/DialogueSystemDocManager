import json
import os
import sys

import openpyxl as op
import xlwings as xw
from openpyxl.styles import Alignment, Font

try:
    import unreal as ue
except ModuleNotFoundError:
    ue = None
from PySide6 import QtCore
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import QApplication, QDialog, QWidget

import window_master as wm


__path__ = os.path.dirname(os.path.realpath(__file__))


class WindowData(object):
    main = None


class WindowModule(object):

    class Main(QWidget):

        def __init__(self, parent=None):

            super(WindowModule.Main, self).__init__(parent)
            self.setAttribute(QtCore.Qt.WA_DeleteOnClose)
            self.setAttribute(QtCore.Qt.WA_QuitOnClose)
            # self.setWindowIcon(WindowData.app_icon)
            # self.setWindowTitle(WindowData.app_title)
            # self.setWindowFlag(
            #     QtCore.Qt.WindowContextHelpButtonHint, False)

            self.ui = None
            self.build_ui(0)

        def build_ui(self, flag):
            ''' build ui by .ui file or .py file
            :arg flag: 
                0 - build by .ui file; 
                1 - build by .py file.
            '''
            if flag == 0:
                ui = "%s/ui/main.ui" % __path__
                self.ui = QUiLoader().load(ui, parentWidget=None)
            elif flag == 1:
                pass
            ui = self.ui


class WindowController(object):

    class Main(object):

        @staticmethod
        def connect():
            ui = WindowData.main.ui
            controller = WindowController.Main
            ui.createBook.clicked.connect(controller.on_createBook_clicked)

        @staticmethod
        def on_createBook_clicked():
            json_path = 'D:/OneDrive/Unreal/PersonalProject/Content/Dialogue'
            json_file = '%s/Dlg_TestFile_2.dlg.json' % json_path
            with open(json_file, encoding="UTF-8") as f:
                file = json.load(f)
                print(file)


class ScriptManager(object):
    @staticmethod
    def list_assets(directory_path):
        '''List all asset in directory.
        :arg directory_path: Long Path Name of the directory.
        '''
        ue.log(f"List all assets in '{directory_path}' directory")
        if not ue.EditorAssetLibrary.does_directory_exist(directory_path):
            ue.log_error("Directory does not exist.")
            return
        if not ue.EditorAssetLibrary.does_directory_have_assets(directory_path):
            ue.log_warning("Directory is empty.")
        else:
            asset_list = ue.EditorAssetLibrary.list_assets(directory_path)
            for asset in asset_list:
                print(asset)


class ExcelData(object):
    work_book = None
    work_sheet = None
    flow = []
    flow_all_paths = []
    json = None
    sheet_row = None
    next_str = ["Next", "Skip", "Finish"]
    head = {
        "GUID": 1,
        "TechIndex": 2,
        "TechJump": 3,
        "WorkIndex": 4,
        "Owner": 5,
        "Content": 6,
        "WorkJump": 13
    }

    class Style(object):
        font = Font(name='Consolas', color='00FF0000')


class ExcelManager(object):
    def __init__(self):
        super(ExcelManager, self).__init__()
        json_path = os.path.abspath(os.path.join(os.getcwd(), "../.."))
        print(json_path)
        json_file = '%s/Dialogue/Dlg_TestFile.dlg.json' % json_path
        self.ExcelWriter.workbook_loader()
        self.ExcelWriter.dlgjson_loader(json_file)
        self.ExcelWriter.start_node_handler()
        print(ExcelData.flow_all_paths)

        self.ExcelWriter.flow_filter()
        # self.ExcelWriter.flow_level_parser()
        self.ExcelWriter.book_writer()
        ExcelData.work_book.save("file/Generate.xlsx")
        print("Gernerated Excel files as 'file\Generate.xlsx'")
        # os.startfile("%s/file/Generate.xlsx" % __path__)

    class ExcelWriter(object):
        @staticmethod
        def workbook_loader(template_file="file/TempDoc.xlsx"):
            # 加载文件
            wb = op.load_workbook(filename=template_file)
            if wb:
                ExcelData.work_book = wb
                ExcelData.work_sheet = wb.active
                print("Template Excel file loaded.")
            else:
                print("Excel Not Load")

        @staticmethod
        def dlgjson_loader(json_file):
            '''read dlg.json file and convert its content to excel cell value'''
            file = None

            with open(json_file, encoding="UTF-8") as f:
                file = json.load(f)
            if not file:
                return

            # 获取开始节点
            ExcelData.json = file
            f.close()

        @staticmethod
        def book_writer():
            flow = ExcelData.flow
            nodes = ExcelData.json["Nodes"]
            start_paser = ExcelManager.ExcelWriter.start_node_parser
            node_parser = ExcelManager.ExcelWriter.speech_node_parser
            child_parser = ExcelManager.ExcelWriter.child_parser
            selector_parser = ExcelManager.ExcelWriter.selector_parser
            sequence_parser = ExcelManager.ExcelWriter.speech_sequence_parser
            end_parser = ExcelManager.ExcelWriter.end_node_parser

            # ExcelManager.ExcelWriter.row_level_parser(flow)

            for node in flow:
                print(node)

            i = 2
            for elm in flow:

                nodeindex = elm["index"]
                elmtype = elm["type"]
                indent = elm["level"] * 2

                if elmtype == "start":
                    row = start_paser(i)
                elif elmtype == "speech":
                    node = nodes[nodeindex]
                    row = node_parser(i, node)
                elif elmtype == "branch":
                    # StartNode's branch.
                    # So I don't know why you want branch after start.
                    # You should consider about SelectorNode.
                    if nodeindex == -1:
                        branch_node = ExcelData.json["StartNode"]
                    # SpeechNode's branch.
                    else:
                        branch_node = nodes[nodeindex]
                    # Get Branch order
                    branch_index = elm['order']
                    row = child_parser(i, branch_node, branch_index)
                elif elmtype == "selector":
                    node = nodes[nodeindex]
                    row = selector_parser(i, node)
                elif elmtype == "sequence":
                    node = nodes[nodeindex]
                    row = sequence_parser(i, node, elmtype)
                elif elmtype == "sequence_child":
                    node = nodes[nodeindex]
                    row = sequence_parser(i, node, elmtype, elm['order'])
                elif elmtype == "sequence_end":
                    node = nodes[nodeindex]
                    row = sequence_parser(i, node, elmtype)
                elif elmtype == "end":
                    node = nodes[nodeindex]
                    row = end_parser(i, node)
                i += 1
                if row:
                    ExcelManager.ExcelStyle.apply_common_style(row["common"])
                    ExcelManager.ExcelStyle.apply_content_style(
                        row["content"], indent)

        @staticmethod
        def start_node_handler():
            start_node = ExcelData.json["StartNode"]
            flp = ExcelData.flow_all_paths
            bhandler = ExcelManager.ExcelWriter.branch_handler

            flow_elm = {}  # 创建一个flow元素
            flow_elm["index"] = -1  # 分支的来源node
            flow_elm["type"] = "start"  # 标记类型
            flow_elm["level"] = 0
            # Add node to flow
            # print(flow_elm)
            flp.append(flow_elm)  # 加入flow
            # 处理分支
            bhandler(start_node)

        @staticmethod
        def branch_handler(node, level=0):
            flp = ExcelData.flow_all_paths
            nhandler = ExcelManager.ExcelWriter.node_handler
            next_str = ExcelData.next_str

            children = node["Children"]

            # Next
            if children.__len__() == 1 and children[0]["Text"] in next_str:
                target_index = children[0]["TargetIndex"]
                nhandler(target_index, level)
                return

            # Branch
            for child in children:
                flow_elm = {}
                try:
                    flow_elm["index"] = node["__index__"]
                except KeyError:
                    flow_elm["index"] = -1
                flow_elm["type"] = "branch"
                flow_elm["order"] = children.index(child)
                flow_elm["level"] = level
                # ExcelManager.ExcelWriter.testtest(flow_elm, level)

                # Add node to flow.
                flp.append(flow_elm)
                # Parse target node
                target_index = child["TargetIndex"]
                nhandler(target_index, level + 1)

        @staticmethod
        def node_handler(node_index=-1, level=0):
            '''Add input node to flow by calling 
            ExcelManager.ExcelWriter.flow_filter().
            It will call itself if the node has braches.
            :param node_index: The index of node to parse.
            '''
            bhandler = ExcelManager.ExcelWriter.branch_handler
            flp = ExcelData.flow_all_paths
            # fl = ExcelManager.ExcelWriter.flow_filter
            nodes: list = ExcelData.json["Nodes"]
            node: dict = nodes[node_index]

            # 节点的出链类型
            ntype = ExcelManager.ExcelWriter.node_type(node)

            flow_elm = {}

            # index
            flow_elm["index"] = node["__index__"]
            # type
            if ntype == "sequence":
                flow_elm["type"] = "sequence"
                flow_elm["level"] = level
                flp.append(flow_elm)

                seq_nodes: list = node["SpeechSequence"]
                for s_node in seq_nodes:
                    flow_elm = {}
                    flow_elm["index"] = node["__index__"]
                    flow_elm["type"] = "sequence_child"
                    flow_elm["order"] = seq_nodes.index(s_node)
                    flow_elm["level"] = level + 1
                    flp.append(flow_elm)

                flow_elm = {}
                flow_elm["type"] = "sequence_end"
            else:
                flow_elm["type"] = ntype
            # level
            flow_elm["level"] = level

            # Add node to flow.
            flp.append(flow_elm)
            bhandler(node, level)

        @staticmethod
        def node_type(node):
            '''Return a node's type.
            :param node: A dialogue system node in json file. Should convert to dict first.
            :type: dict
            :returns:
                :return "start": It's a start node.
                :return "speech": It's a speech node.
                :return "selector": It's a selector node.
                :return "sequence": It's a sequence node.
                :return "end": It's a end node.
            :rtype: str
            '''
            if node["__type__"] == "DlgNode_Speech":
                if "__index__" not in node.keys():
                    return "start"
                else:
                    return "speech"
            elif node["__type__"] == "DlgNode_Selector":
                return "selector"
            elif node["__type__"] == "DlgNode_SpeechSequence":
                return "sequence"
            elif node["__type__"] == "DlgNode_End":
                return "end"

        @staticmethod
        def flow_filter():
            """Write element to flow list, and pop the existed same element.
            :param flow_element: New flow element from flow parser.
            """
            '''
            找到每个节点的入链数量，
            如果单个节点的入链数 n >= 2
            回溯来源节点，
            比较所有level，找到最小的level
            '''
            '''
            其实相当于找到每个节点在flow中的出现次数，
            如果出现次数大于等于2，则找到最高的合并等级，
            向前合并
            '''

        @staticmethod
        def start_node_parser(row):
            ws = ExcelData.work_sheet
            head = ExcelData.head
            start_node = ExcelData.json["StartNode"]
            # 处理开始节点的基础数据
            c_guid = ws.cell(row, head["GUID"], start_node["NodeGUID"])
            c_tech_index = ws.cell(row, head["TechIndex"], "Start")
            c_work_index = ws.cell(row, head["WorkIndex"], "Start")
            c_owner = ws.cell(row, head["Owner"], "")
            c_content = ws.cell(row, head["Content"], "【对话开始】")
            if start_node["Children"].__len__() >= 2:
                c_tech_jump = ws.cell(row, head["TechJump"], "Start_Branch")
            else:
                c_tech_jump = ws.cell(row, head["TechJump"],
                                      start_node["Children"][0]["TargetIndex"])

            row = {}
            row["common"] = [c_guid, c_tech_index,
                             c_work_index, c_owner, c_tech_jump]
            row["content"] = [c_content]

            return row

        @staticmethod
        def speech_node_parser(row, node):
            ws = ExcelData.work_sheet
            head = ExcelData.head

            c_guid = ws.cell(row, head["GUID"], node["NodeGUID"])
            c_tech_index = ws.cell(row, head["TechIndex"], node["__index__"])
            c_owner = ws.cell(row, head["Owner"], node["OwnerName"])
            c_content = ws.cell(row, head["Content"], node["Text"])
            if node["Children"].__len__() >= 2:
                c_tech_jump = ws.cell(row, head["TechJump"], str(
                    node["__index__"]) + "_Branch")
            else:
                target_index = node["Children"][0]["TargetIndex"]
                c_tech_jump = ws.cell(row, head["TechJump"], target_index)

            row = {}
            row["common"] = [c_guid, c_tech_index, c_owner, c_tech_jump]
            row["content"] = [c_content]

            return row

        @staticmethod
        def child_parser(row, node, child_index):
            ws = ExcelData.work_sheet
            head = ExcelData.head
            children = node["Children"]
            child = children[child_index]
            # TechIndex
            try:
                tech_index = node["__index__"]
            except KeyError:
                # StartNode does not have "__index__" attribute.
                tech_index = -1
            c_tech_index = ws.cell(row, head["TechIndex"], tech_index)
            # TechJump
            c_tech_jump = ws.cell(row, head["TechJump"], child["TargetIndex"])
            c_work_index = ws.cell(row, head["WorkIndex"], tech_index)
            c_content = ws.cell(row, head["Content"], ">> " + child["Text"])

            row = {}
            row["common"] = [c_tech_index, c_work_index, c_tech_jump]
            row["content"] = [c_content]

            return row

        @staticmethod
        def selector_parser(row, node):
            ws = ExcelData.work_sheet
            head = ExcelData.head

            c_guid = ws.cell(row, head["GUID"], node["NodeGUID"])
            c_tech_index = ws.cell(row, head["TechIndex"], node["__index__"])
            c_owner = ws.cell(row, head["Owner"], node["OwnerName"])
            sel_type: str = node["SelectorType"]
            sel_type = sel_type.split("::")[-1]
            c_content = ws.cell(
                row, head["Content"], f"> Selector [{sel_type}]")
            c_tech_jump = ws.cell(row, head["TechJump"], str(
                node["__index__"]) + "_Select")
            row = {}
            row["common"] = [c_guid, c_tech_index, c_owner, c_tech_jump]
            row["content"] = [c_content]

            return row

        @staticmethod
        def speech_sequence_parser(row, node, seq_type, seq_index=-1):
            ws = ExcelData.work_sheet
            head = ExcelData.head
            if seq_type == "sequence":
                c_guid = ws.cell(row, head["GUID"], node["NodeGUID"])
                c_tech_index = ws.cell(
                    row, head["TechIndex"], node["__index__"])
                c_owner = ws.cell(row, head["Owner"], node["OwnerName"])
                c_content = ws.cell(row, head["Content"], "## 垂直序列")
                c_tech_jump = ws.cell(row, head["TechJump"],
                                      node["__index__"])

                row = {}
                row["common"] = [c_guid, c_owner, c_tech_index, c_tech_jump]
                row["content"] = [c_content]

                return row

            elif seq_type == "sequence_child":
                child_node = node["SpeechSequence"][seq_index]

                c_tech_index = ws.cell(
                    row, head["TechIndex"], node["__index__"])
                c_owner = ws.cell(row, head["Owner"], child_node["Speaker"])
                c_content = ws.cell(row, head["Content"], child_node["Text"])
                c_tech_jump = ws.cell(row, head["TechJump"], str(
                    node["InnerEdges"][seq_index]["TargetIndex"]))

                row = {}
                row["common"] = [c_owner, c_tech_index, c_tech_jump]
                row["content"] = [c_content]

                return row

            elif seq_type == "sequence_end":
                c_tech_index = ws.cell(
                    row, head["TechIndex"], node["__index__"])
                c_owner = ws.cell(row, head["Owner"], "Sequence")
                c_content = ws.cell(row, head["Content"], "## 垂直序列结束")
                c_tech_jump = ws.cell(row, head["TechJump"], str(
                    node["Children"][0]["TargetIndex"]))

                row = {}
                row["common"] = [c_owner, c_tech_index, c_tech_jump]
                row["content"] = [c_content]

                return row

        @staticmethod
        def end_node_parser(row, node):
            ws = ExcelData.work_sheet
            head = ExcelData.head
            # Write data.
            c_guid = ws.cell(row, head["GUID"], node["NodeGUID"])
            c_tech_index = ws.cell(row, head["TechIndex"], node["__index__"])
            c_owner = ws.cell(row, head["Owner"], node["OwnerName"])
            c_work_index = ws.cell(row, head["WorkIndex"], "End")
            c_content = ws.cell(row, head["Content"], "【对话结束】")
            c_tech_jump = ws.cell(row, head["TechJump"], "End")

            row = {}
            row["common"] = [c_guid, c_tech_index,
                             c_owner, c_work_index, c_tech_jump]
            row["content"] = [c_content]

            return row

    class ExcelStyle(object):
        @staticmethod
        def apply_common_style(cell_list):
            for cell in cell_list:
                cell.font = Font(name="Consolas")
                cell.alignment = Alignment(horizontal="left")

        @staticmethod
        def apply_content_style(cell_list, indent=0):
            for cell in cell_list:
                cell.font = Font(name='微软雅黑')
                cell.alignment = Alignment(horizontal="left", indent=indent)


class WindowManager(object):

    @staticmethod
    def create_window_main():
        wm.window_delete(WindowData.main)
        WindowData.main = WindowModule.Main()
        wm.show_window(WindowData.main.ui)

    @staticmethod
    def create_window_main_without_ue():
        WindowData.main = WindowModule.Main()
        WindowData.main.ui.show()


if __name__ == '__main__':
    # Create APP
    app = None
    if not QApplication.instance():
        app = QApplication(sys.argv)
    else:
        app = QApplication.instance()
    app.exit()

    # Create window
    if ue:
        # Create window in UE
        ue.log("Load Story Doc Manager Window")
        WindowManager.create_window_main()
        WindowController.Main.connect()
    else:
        # Create window without UE
        print("Start Story Doc Manager without Unreal Engine")
        # WindowManager.create_window_main_without_ue()
        # WindowController.Main.connect()
        ExcelManager()
        # sys.exit(app.exec())


def __exit__():
    wm.window_delete(WindowData.main)
